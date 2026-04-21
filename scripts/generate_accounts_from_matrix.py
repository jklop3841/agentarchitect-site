#!/usr/bin/env python3
"""
Generate a sanitized social-publishing account config from the user's Excel matrix.

The output intentionally excludes passwords, cookies, browser state, and free-form
password-hint notes. Phone/login identifiers are retained because the workflow
needs to map known accounts to platform targets.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MATRIX = Path(r"C:\Users\97566\Desktop\自媒体平台账号矩阵管理表.xlsx")
DEFAULT_OUTPUT = ROOT / "docs" / "social-publishing" / "accounts.generated.json"
FIRST_STAGE = {"toutiao", "baijiahao", "zhihu", "sohu", "csdn", "juejin"}
MANUAL_STAGE = {"wechat"}

PLATFORM_IDS = {
    "微信公众号": "wechat",
    "头条号": "toutiao",
    "百家号": "baijiahao",
    "知乎": "zhihu",
    "企鹅号": "qqmedia",
    "微博": "weibo",
    "搜狐号": "sohu",
    "CSDN": "csdn",
    "腾讯开发者社区": "tencentcloud",
    "阿里云开发者社区": "aliyun",
    "稀土掘金": "juejin",
    "sengmentfault": "segmentfault",
    "SegmentFault": "segmentfault",
    "博客园": "cnblogs",
    "豆丁网": "docin",
    "电子发烧友": "elecfans",
}

VARIANTS = {
    "csdn": "technical",
    "juejin": "technical",
    "tencentcloud": "technical",
    "aliyun": "technical",
    "segmentfault": "technical",
    "cnblogs": "technical",
    "toutiao": "general",
    "baijiahao": "general",
    "zhihu": "general",
    "sohu": "general",
    "qqmedia": "general",
    "weibo": "general",
    "wechat": "manual",
}

SENSITIVE_QUERY_KEYS = {"token", "access_token", "refresh_token", "secret", "key", "ticket", "signature"}


def normalize_login(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def sanitize_url(value: Any) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    parts = urlsplit(raw)
    safe_query = [
        (key, val)
        for key, val in parse_qsl(parts.query, keep_blank_values=True)
        if key.lower() not in SENSITIVE_QUERY_KEYS
    ]
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(safe_query), parts.fragment))


def stage_for(platform_id: str) -> str:
    if platform_id in FIRST_STAGE:
        return "first-stage"
    if platform_id in MANUAL_STAGE:
        return "manual"
    return "configured"


def mode_for(platform_id: str) -> str:
    return "manual" if platform_id in MANUAL_STAGE else "draft"


def adapter_for(platform_id: str) -> str:
    if platform_id in FIRST_STAGE:
        return "wechatsync"
    if platform_id in MANUAL_STAGE:
        return "manual"
    return "future-adapter"


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate sanitized social publishing account config")
    parser.add_argument("--input", type=Path, default=DEFAULT_MATRIX)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    workbook = load_workbook(args.input, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        platform_name = str(record.get("平台名称") or "").strip()
        if not platform_name:
            continue
        platform_id = PLATFORM_IDS.get(platform_name, platform_name.lower())
        rows.append({
            "platform": platform_id,
            "platformName": platform_name,
            "accountNickname": record.get("账号昵称"),
            "loginIdentifier": normalize_login(record.get("登录账号/绑定手机号")),
            "backendUrl": sanitize_url(record.get("官方后台地址")),
            "geoTarget": record.get("GEO对象（态射）"),
            "adapter": adapter_for(platform_id),
            "stage": stage_for(platform_id),
            "defaultMode": mode_for(platform_id),
            "variant": VARIANTS.get(platform_id, "general"),
        })

    output = {
        "source": str(args.input),
        "generatedBy": "scripts/generate_accounts_from_matrix.py",
        "credentialPolicy": "No passwords, cookies, storageState, browser profiles, or token values are stored here.",
        "firstStagePlatforms": sorted(FIRST_STAGE),
        "accounts": rows,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"output": str(args.output), "accounts": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
